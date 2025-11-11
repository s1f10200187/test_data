#!/usr/bin/env python3
"""
Excel to Markdown Converter

このスクリプトは、指定されたディレクトリ配下の全てのExcelファイル(.xlsx)を
markitdownライブラリを使用してMarkdown形式に変換します。
画像が含まれるシートがある場合、画像を抽出してファイルとして保存し、
各シートの末尾に画像を追加します。

画像の保存場所:
    - Excelファイルと同じ階層に「{Excelファイル名}_images」ディレクトリを作成
    - 抽出された画像はこのディレクトリに保存されます
    - Markdownファイル内では相対パスで参照されます

使用方法:
    python excel_to_markdown.py <directory_path> [options]

オプション:
    --dry-run
        実際には変換せず、変換対象ファイルのみを表示します。

    --output-dir <path>
        Markdownファイルの出力先ディレクトリを指定します。
        指定しない場合、Excelファイルと同じディレクトリに出力されます。
        ※画像ディレクトリは常にExcelファイルと同じ階層に作成されます。

    --llm-images
        LLM（OpenAI）を使用して画像の説明を生成します。
        環境変数 OPENAI_API_KEY が必要です。

使用例:
    # 基本的な変換
    python excel_to_markdown.py ./sample_data

    # 変換対象ファイルの確認のみ
    python excel_to_markdown.py ./sample_data --dry-run

    # 出力先を指定
    python excel_to_markdown.py ./sample_data --output-dir ./markdown_output

    # LLMで画像説明を生成
    OPENAI_API_KEY=xxx python excel_to_markdown.py ./sample_data --llm-images

出力例:
    sample_data/
    ├── data.xlsx
    ├── data.md              # 生成されたMarkdown
    └── data_images/         # 画像格納ディレクトリ
        ├── Sheet1_image_1.png
        └── Sheet1_image_2.jpg

必須条件:
    pip install markitdown openpyxl pillow
"""

import argparse
import io
import sys
from pathlib import Path
from typing import List, Dict, Optional

try:
    from markitdown import MarkItDown
    from openpyxl import load_workbook
    from PIL import Image
except ImportError as e:
    print(f"必要なライブラリがインストールされていません: {e}", file=sys.stderr)
    print("以下のコマンドでインストールしてください:", file=sys.stderr)
    print("pip install markitdown openpyxl pillow", file=sys.stderr)
    sys.exit(1)


def find_excel_files(directory: Path) -> List[Path]:
    """
    指定されたディレクトリ配下の全てのExcelファイルを再帰的に検索します。

    Args:
        directory: 検索対象のディレクトリパス

    Returns:
        Excelファイルのパスリスト
    """
    excel_files = []

    # .xlsx ファイルを再帰的に検索
    for excel_file in directory.rglob("*.xlsx"):
        # 一時ファイル（~$で始まるファイル）を除外
        if not excel_file.name.startswith("~$"):
            excel_files.append(excel_file)

    return sorted(excel_files)


def extract_images_from_excel(excel_path: Path, image_dir: Path) -> Dict[str, List[str]]:
    """
    Excelファイルから画像を抽出してファイルとして保存します。

    Args:
        excel_path: Excelファイルのパス
        image_dir: 画像の保存先ディレクトリ

    Returns:
        シート名をキーとし、画像ファイルパスのリストを値とする辞書
    """
    images_by_sheet = {}

    try:
        wb = load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_images = []

            # シート内の画像を取得
            if hasattr(sheet, '_images') and sheet._images:
                for img_idx, img in enumerate(sheet._images, 1):
                    try:
                        # 画像データを取得
                        if hasattr(img, '_data'):
                            image_data = img._data()
                        elif hasattr(img, 'ref'):
                            # 別の方法で画像データを取得
                            image_data = img.ref
                        else:
                            continue

                        # PIL Imageに変換
                        pil_image = Image.open(io.BytesIO(image_data))
                        image_format = pil_image.format or 'PNG'

                        # ファイル拡張子を決定
                        ext = image_format.lower()
                        if ext == 'jpeg':
                            ext = 'jpg'

                        # 画像ファイル名を生成（シート名_画像番号.拡張子）
                        # ファイル名に使えない文字を置換
                        safe_sheet_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in sheet_name)
                        image_filename = f"{safe_sheet_name}_image_{img_idx}.{ext}"
                        image_path = image_dir / image_filename

                        # 画像を保存
                        pil_image.save(image_path, format=image_format)
                        sheet_images.append(image_filename)

                    except Exception as e:
                        print(f"  警告: 画像の抽出に失敗しました: {e}", file=sys.stderr)
                        continue

            if sheet_images:
                images_by_sheet[sheet_name] = sheet_images

        wb.close()
    except Exception as e:
        print(f"  警告: Excelファイルの画像抽出中にエラーが発生しました: {e}", file=sys.stderr)

    return images_by_sheet


def convert_excel_to_markdown(
    excel_path: Path,
    output_dir: Optional[Path] = None,
    llm_client=None
) -> Optional[Path]:
    """
    Excelファイルを画像付きMarkdownに変換します。
    画像はExcelファイルと同じ階層に作成される画像ディレクトリに保存されます。

    Args:
        excel_path: Excelファイルのパス
        output_dir: 出力先ディレクトリ（Noneの場合、元のファイルと同じディレクトリ）
        llm_client: LLMクライアント（画像説明生成用、オプション）

    Returns:
        生成されたMarkdownファイルのパス、失敗時はNone
    """
    try:
        # 画像保存用ディレクトリを作成（Excelファイルと同じ階層）
        image_dir_name = f"{excel_path.stem}_images"
        image_dir = excel_path.parent / image_dir_name

        # 画像を抽出（画像が存在する場合のみディレクトリを作成）
        images_by_sheet = {}
        wb = load_workbook(excel_path, data_only=True)
        has_images = any(
            hasattr(wb[sheet_name], '_images') and wb[sheet_name]._images
            for sheet_name in wb.sheetnames
        )
        wb.close()

        if has_images:
            image_dir.mkdir(exist_ok=True)
            images_by_sheet = extract_images_from_excel(excel_path, image_dir)

        # markitdownで基本的な変換
        if llm_client:
            md_converter = MarkItDown(llm_client=llm_client, llm_model="gpt-4o")
        else:
            md_converter = MarkItDown()

        result = md_converter.convert(str(excel_path))

        # Markdownコンテンツを取得してシートごとに分割・再構築
        markdown_content = result.text_content

        # 画像が存在する場合、各シートの末尾に画像を追加
        if images_by_sheet:
            # シートごとに画像を末尾に追加
            for sheet_name, image_files in images_by_sheet.items():
                # シート名をMarkdownから検索して、その末尾に画像を追加
                sheet_header = f"## {sheet_name}"

                if sheet_header in markdown_content:
                    # 次のシートの開始位置を探す
                    lines = markdown_content.split('\n')
                    new_lines = []
                    in_target_sheet = False
                    images_added = False

                    for i, line in enumerate(lines):
                        new_lines.append(line)

                        # 対象シートの開始
                        if line.strip() == sheet_header:
                            in_target_sheet = True
                        # 次のシートの開始、またはファイル末尾で画像を挿入
                        elif in_target_sheet and (line.startswith('## ') or i == len(lines) - 1):
                            if not images_added:
                                # 画像を挿入
                                if i == len(lines) - 1 and not line.startswith('## '):
                                    # 最後のシートの場合、現在の行の後に追加
                                    pass
                                else:
                                    # 次のシートの前に追加（現在の行の前に挿入するため一旦削除）
                                    new_lines.pop()

                                new_lines.append('')
                                new_lines.append('### 画像')
                                new_lines.append('')

                                for img_file in image_files:
                                    img_path = f"{image_dir_name}/{img_file}"
                                    new_lines.append(f"![{img_file}]({img_path})")
                                    new_lines.append('')

                                if i != len(lines) - 1:
                                    new_lines.append(line)  # 次のシートの開始行を再追加

                                images_added = True
                                in_target_sheet = False

                    markdown_content = '\n'.join(new_lines)
                else:
                    # シートヘッダーが見つからない場合は末尾に追加
                    markdown_content += f"\n\n## {sheet_name}\n\n### 画像\n\n"
                    for img_file in image_files:
                        img_path = f"{image_dir_name}/{img_file}"
                        markdown_content += f"![{img_file}]({img_path})\n\n"

        # 出力先を決定
        if output_dir:
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"{excel_path.stem}.md"
        else:
            output_path = excel_path.with_suffix('.md')

        # Markdownファイルを保存
        output_path.write_text(markdown_content, encoding='utf-8')

        return output_path

    except Exception as e:
        print(f"  エラー: 変換に失敗しました: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return None


def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(
        description='Excelファイルを画像埋め込み付きMarkdownに一括変換します。',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )

    parser.add_argument(
        'directory',
        type=Path,
        help='変換対象のExcelファイルが含まれるディレクトリパス'
    )

    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='変換対象ファイルを表示するのみで実際の変換は行わない'
    )

    parser.add_argument(
        '--output-dir',
        type=Path,
        help='Markdownファイルの出力先ディレクトリ'
    )

    parser.add_argument(
        '--llm-images',
        action='store_true',
        help='LLM（OpenAI）を使用して画像の説明を生成（OPENAI_API_KEY環境変数が必要）'
    )

    args = parser.parse_args()

    # ディレクトリの存在確認
    if not args.directory.exists():
        print(f"エラー: ディレクトリが見つかりません: {args.directory}", file=sys.stderr)
        sys.exit(1)

    if not args.directory.is_dir():
        print(f"エラー: 指定されたパスはディレクトリではありません: {args.directory}", file=sys.stderr)
        sys.exit(1)

    # LLMクライアントの設定
    llm_client = None
    if args.llm_images:
        try:
            from openai import OpenAI
            llm_client = OpenAI()
            print("LLMによる画像説明生成が有効です。")
        except ImportError:
            print("警告: openaiライブラリがインストールされていません。", file=sys.stderr)
            print("pip install openai でインストールしてください。", file=sys.stderr)
        except Exception as e:
            print(f"警告: OpenAIクライアントの初期化に失敗しました: {e}", file=sys.stderr)

    # Excelファイルを検索
    print(f"ディレクトリを検索中: {args.directory}")
    excel_files = find_excel_files(args.directory)

    if not excel_files:
        print("変換対象のExcelファイルが見つかりませんでした。")
        return

    print(f"\n見つかったExcelファイル: {len(excel_files)}件")
    for excel_file in excel_files:
        print(f"  - {excel_file}")

    # dry-runモードの場合はここで終了
    if args.dry_run:
        print("\n--dry-run モードのため、変換は実行されませんでした。")
        return

    # 変換処理
    print("\n変換を開始します...")
    successful = 0
    failed = 0

    for excel_file in excel_files:
        print(f"\n変換中: {excel_file.name}")

        output_path = convert_excel_to_markdown(
            excel_file,
            args.output_dir,
            llm_client
        )

        if output_path:
            print(f"  ✓ 成功: {output_path}")
            successful += 1
        else:
            print(f"  ✗ 失敗")
            failed += 1

    # 結果サマリー
    print("\n" + "="*50)
    print("変換完了")
    print(f"成功: {successful}件")
    print(f"失敗: {failed}件")
    print("="*50)


if __name__ == "__main__":
    main()
